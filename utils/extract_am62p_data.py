#!/usr/bin/env python3
"""
Extract AM62P Power Estimation Tool data from XLSM file
This script parses the AM62P Excel file and extracts all fields, calculations, and data structures
"""

import openpyxl
import json
import sys
from pathlib import Path

def extract_am62p_data(xlsm_path):
    """Extract all relevant data from AM62P Power Estimation Tool"""

    print(f"Loading workbook: {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)

    # Get all sheet names
    print(f"\nAvailable sheets: {wb.sheetnames}")

    data = {
        "id": "am62p",
        "name": "AM62P",
        "description": "Multi-core Arm Cortex-A53 processor with integrated vision acceleration",
        "sheets": {},
        "inputs": {},
        "calculations": {},
        "dropdowns": {},
        "use_cases": []
    }

    # Extract data from each relevant sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*60}")
        print(f"Processing sheet: {sheet_name}")
        print(f"{'='*60}")

        ws = wb[sheet_name]
        sheet_data = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "cells": {}
        }

        # Extract all non-empty cells with their values and formulas
        cell_count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200),
                                 min_col=1, max_col=min(ws.max_column, 50)):
            for cell in row:
                if cell.value is not None:
                    cell_ref = cell.coordinate
                    cell_info = {
                        "value": str(cell.value),
                        "data_type": str(cell.data_type)
                    }

                    # Extract formula if present
                    if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_info["formula"] = cell.value

                    # Check for data validation (dropdowns)
                    if cell.data_validation and hasattr(cell.data_validation, 'formula1'):
                        cell_info["validation"] = str(cell.data_validation.formula1)

                    sheet_data["cells"][cell_ref] = cell_info
                    cell_count += 1

        print(f"  Extracted {cell_count} non-empty cells")
        data["sheets"][sheet_name] = sheet_data

    return data

def identify_key_sections(data):
    """Identify key sections in the workbook"""

    print("\n" + "="*60)
    print("IDENTIFYING KEY SECTIONS")
    print("="*60)

    key_patterns = [
        "temperature", "temp", "junction",
        "voltage", "vdd", "supply",
        "frequency", "opp", "mhz",
        "utilization", "util", "percent",
        "ddr", "memory",
        "usb", "mmc", "sd", "ethernet",
        "power", "leakage", "dynamic", "static",
        "rail", "total"
    ]

    findings = {}

    for sheet_name, sheet_data in data["sheets"].items():
        sheet_findings = []
        for cell_ref, cell_info in sheet_data["cells"].items():
            value_lower = cell_info["value"].lower()
            for pattern in key_patterns:
                if pattern in value_lower:
                    sheet_findings.append({
                        "cell": cell_ref,
                        "value": cell_info["value"],
                        "pattern": pattern
                    })
                    break

        if sheet_findings:
            findings[sheet_name] = sheet_findings
            print(f"\n{sheet_name}: Found {len(sheet_findings)} key cells")
            for finding in sheet_findings[:5]:  # Show first 5
                print(f"  {finding['cell']}: {finding['value']} (matched: {finding['pattern']})")

    return findings

def main():
    xlsm_path = Path(__file__).parent.parent / "SPRUJD9_AM62P_PET_1_1.xlsm"

    if not xlsm_path.exists():
        print(f"Error: File not found: {xlsm_path}")
        sys.exit(1)

    # Extract data
    data = extract_am62p_data(xlsm_path)

    # Identify key sections
    findings = identify_key_sections(data)

    # Save to JSON
    output_path = Path(__file__).parent / "am62p_extracted.json"
    print(f"\n{'='*60}")
    print(f"Saving extracted data to: {output_path}")
    print(f"{'='*60}")

    with open(output_path, 'w') as f:
        json.dump(data, f, indent=2)

    # Save findings separately
    findings_path = Path(__file__).parent / "am62p_findings.json"
    with open(findings_path, 'w') as f:
        json.dump(findings, f, indent=2)

    print(f"\nExtraction complete!")
    print(f"  Data saved to: {output_path}")
    print(f"  Findings saved to: {findings_path}")
    print(f"\nTotal sheets processed: {len(data['sheets'])}")

    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    for sheet_name, sheet_data in data["sheets"].items():
        print(f"{sheet_name}: {len(sheet_data['cells'])} cells")

if __name__ == "__main__":
    main()
