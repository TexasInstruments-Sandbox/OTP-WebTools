import openpyxl
import json
import os

EXCEL_FILE = '../AM62x_Power_Estimation_Tool_Public_1v1.xlsm'

def extract_mapping():
    print("Loading workbook to parse formulas...")
    # data_only=False allows reading the raw formulas (e.g., =MPU_A530_Util)
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    ws_use = wb['Use Cases']

    # Locate the 'Public PET UC' block
    col_util = None
    col_mode = None
    for col in range(1, ws_use.max_column + 1):
        if ws_use.cell(row=3, column=col).value == 'Public PET UC':
            col_util = col
            col_mode = col + 1
            break

    if not col_util:
        print("Error: Could not find 'Public PET UC' column.")
        return

    mapping = {}

    for row in range(5, ws_use.max_row + 1):
        phys = ws_use.cell(row=row, column=4).value # D
        inst = ws_use.cell(row=row, column=3).value # C
        
        if not phys:
            continue
            
        key = f"{str(phys).strip()}_{str(inst).strip()}"
        
        val_util = ws_use.cell(row=row, column=col_util).value
        val_mode = ws_use.cell(row=row, column=col_mode).value
        
        def clean_formula(val):
            if not isinstance(val, str):
                return None
            if not val.startswith('='):
                return None
            # Remove =, +, spaces, and trailing noise if any
            clean = val.replace('=', '').replace('+', '').strip()
            # If it's a fixed value formula like "=1" or ="0.01", skip
            try:
                float(clean)
                return None
            except ValueError:
                pass
            
            # Special case for Summary sheet direct refs -> "Summary!C38"
            if 'Summary!' in clean:
                return clean
                
            # Special case for IF formulas -> =IF(Summary!G38="Max", "Strong", "Nominal")
            if clean.startswith('IF('):
                return clean
                
            return clean

        util_ref = clean_formula(val_util)
        mode_ref = clean_formula(val_mode)

        mapping[key] = {
            "utilization_ref": util_ref,
            "mode_ref": mode_ref,
            # Also read the raw value (if possible using data_only=True in another pass, but for now we default to None)
        }

    # Now we need actual default values for these inputs. 
    # The Summary sheet defines the defaults.
    # Instead of parsing everything, the easiest way is to use pandas to get the data_only evaluating values
    print("Exporting to am62x_pet_mapping.js...")
    with open('../data/am62x_pet_mapping.js', 'w') as f:
        f.write("window.TI_AM62X_PET_MAPPING = " + json.dumps(mapping, indent=4) + ";\n")
    print(f"Exported mapping for {len(mapping)} subchips.")

if __name__ == "__main__":
    extract_mapping()
